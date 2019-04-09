import exitroutines
# exit and cleanup

import csv
# needed to read raw data in CSV

import os
import glob
import sys
# needed for directory manipulation

import configparser
# for config file

import numpy as np
# numerical functions

import io
# needed for separating bitstreams

import openpyxl
# needed to save to XLSX

import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
# plotting

# Initialization of parameters1
config = configparser.ConfigParser()

#counter variables
warning = 0
verbose = 10

if os.path.exists('config.ini'):
    config.read('config.ini')
else:
    print('WARNING: Configuration file not found. Initiating config file with default values')
    warning += 1
    configfile = open('config.ini', 'w')
    config['Directories'] = {'WorkingDir' : os.path.dirname(os.path.realpath(__file__))}
    config['Settings'] = {'Verbose' : 10}
    config.write(configfile)
    configfile.flush()
    configfile.close()
    print('Done')

if 'WorkingDir' in config['Directories']:
    workdirectory = config['Directories']['WorkingDir']
else:
    print('ERROR: Working directory not in configuration. Is the ini file correct and not corrupted?')
    exitroutines.getout(warning, -1)

try:
    verbose = (int)(config['Settings']['Verbose'])
except:
    print('WARNING: Verbose not in configuration file. Assigning default value: ', verbose)
    warning += 1

print('Changing to working directory...')
try:
    os.chdir(os.path.join(workdirectory, 'data'))
except:
    print('ERROR: Could not change to working directory. Does directory exist?')
    exitroutines.getout(warning, -1)

numberofcsvs = [i for i in glob.glob('*.{}'.format('csv'))]

if numberofcsvs == []:
    print('Nothing to do, no CSV files found')
    exitroutines.getout(warning, 1)

maindata=[]
voltages=[]
currents=[]
charges=[]
resistances=[]
fluxes=[]
conductances=[]
times=[]

currentflux = 0
currentcharge = 0

for i in numberofcsvs:
    data = np.genfromtxt(i, skip_header=3, skip_footer=0, names=['Time', 'Voltage', 'Current'], delimiter=",")
    maindata.append(data)

wb = openpyxl.Workbook()
sheet = wb.active

for j in range(0,len(maindata[0]['Voltage'])):
               meanV = 0
               meanI = 0
               meanT = 0
               for l in range(0,len(maindata)):
                   meanI = meanI + maindata[l]['Current'][j]
                   meanV = meanV + maindata[l]['Voltage'][j]
                   meanT = meanT + maindata[l]['Time'][j]
               meanV = meanV/len(maindata)
               meanI = meanI/len(maindata)
               meanT = meanT/len(maindata)
               voltages.append(meanV)
               currents.append(meanI)
               times.append(meanT)

               resistances.append(meanV/meanI)
               conductances.append(meanI/meanV)

               c1 = sheet.cell(row = j+1, column = 1)
               c1.value = meanT
               c1 = sheet.cell(row = j+1, column = 2)
               c1.value = meanV
               c1 = sheet.cell(row = j+1, column = 3)
               c1.value = meanI



wb.save(os.path.join(workdirectory, 'results.xlsx')) 
